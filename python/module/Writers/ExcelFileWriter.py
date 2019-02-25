# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import PatternFill


class ExcelWriterBase:

    def __init__(self, filename, source):
        self.filename = self.__CleanFileName__(filename)
        self.source = source
        
    def __CleanFileName__(self, filename):
        for trash in [
                '.do', '.txt', '.dat', '.dta',
                '_const', '_val', '_var', '_validate', '_rename', '.'
                ]:
            filename = str(filename).replace(trash, '')
        
        return filename
    
    def __SetFileNameTag__(self):
        self.tag = ''

    def __OpenFile__(self):
        pass

    def __WriteHeader__(self):
        pass

    def __WriteMainPart__(self):
        pass

    def __WriteFooter__(self):
        self.wb.save(self.filename + self.tag + '.xlsx')
        
    def __PrintEndMessage__(self):
        pass

    def WriteExcelFile(self):
        self.__SetFileNameTag__()
        self.__OpenFile__()
        self.__WriteHeader__()
        self.__WriteMainPart__()
        self.__WriteFooter__()
        self.__PrintEndMessage__()


class CleanLayoutWriter(ExcelWriterBase):
            
    def __SetFileNameTag__(self):
        self.tag = '_layout'
    
    def __OpenFile__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        
    def __WriteHeader__(self):
        header = ['Description', 'Varname', 'Begin' , 'End', 'Value']
        for c, word in enumerate(header):
            self.ws.cell(row=1, column=c+1, value=word)
        
    def __GetValues__(self, var):
        vals = ''
        for val, lab in zip(var.val_list, var.val_label_list):
            vals = vals + str(val) + ': ' + str(lab) + '; '
        
        return vals
            
    def __WriteMainPart__(self):
        for r, var in enumerate(self.source):
            self.ws.cell(row=r+2, column=1, value=str(var.GetFullDescription()))
            self.ws.cell(row=r+2, column=2, value=str(var.name))
            self.ws.cell(row=r+2, column=3, value=str(int(float(var.pos_s))))
            self.ws.cell(row=r+2, column=4, value=str(int(float(var.pos_e))))
            self.ws.cell(row=r+2, column=5, value=str(self.__GetValues__(var)))


class RenameExcelWriter(ExcelWriterBase):

    def __OpenFile__(self):
        self.wb = openpyxl.Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = 'Description'
        self.ws2 = self.wb.create_sheet(title='VarName')
    
    def __GetNumFiles__(self):
        first_key = sorted(self.source.keys())[0]
        return len(self.source[first_key].list)
        
    def __WriteHeader__(self):
        self.row = 0
        self.ws1.cell(row=self.row+1, column=1, value='Description: Base')
        self.ws2.cell(row=self.row+1, column=1, value='VarName: Base')
        for c in range(self.__GetNumFiles__()):
            self.ws1.cell(self.row+1, column=c+2, value='Description: InFile ' + str(c))
            self.ws2.cell(self.row+1, column=c+2, value='VarName: InFile ' + str(c))

    def __WriteMainPart__(self):
        fill = PatternFill(patternType='solid', fgColor='ffc040')
        for key, synonym in self.source.items():
            self.row = self.row + 1
            self.ws1.cell(row=self.row+1, column=1, value=synonym.baseinfo.GetFullDescription())
            self.ws2.cell(row=self.row+1, column=1, value=synonym.baseinfo.name)
            for c, var in enumerate(synonym.list):
                try:
                    self.ws1.cell(row=self.row+1, column=c+2, value=var.GetFullDescription())
                    self.ws2.cell(row=self.row+1, column=c+2, value=var.name)

                    # Put color to cells 
                    if c != 0 and var.GetFullDescription() != synonym.list[c-1].GetFullDescription():
                        for ws in [self.ws1, self.ws2]:
                            ws.cell(row=self.row+1, column=c+1).fill = fill
                            ws.cell(row=self.row+1, column=c+2).fill = fill
                except (IndexError, AttributeError):
                    pass
                        

    def __PrintEndMessage__(self):
        print('Rename table (.xlsx): Done')